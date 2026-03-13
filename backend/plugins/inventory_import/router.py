from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import json
from typing import List, Optional

from app.db.database import get_db
from app.db.models import (
    InventoryItem, InventoryCategory, Location, ItemLocationStock, 
    InventoryAuditLog, AuditLogAction, Manufacturer, Supplier
)
from app.core.dependencies import require_admin_access

router = APIRouter(prefix="/plugins/inventory-import", tags=["plugin-inventory-import"])

# --- POMOCNÉ FUNKCE ---

async def resolve_category_path(db: AsyncSession, company_id: int, path_str: str) -> InventoryCategory:
    """
    Zpracuje textový řetězec cesty, např. "Elektro > Kabely > Silové".
    Postupně hledá nebo vytváří kategorie v hierarchii.
    Vrací poslední kategorii v řetězci (tu, která se přiřadí položce).
    """
    # Rozdělit podle '>' a ořezat mezery
    parts = [p.strip() for p in path_str.split('>')]
    
    current_parent_id = None
    last_category = None

    for name in parts:
        if not name:
            continue
            
        # Hledáme kategorii s tímto názvem A tímto rodičem
        stmt = select(InventoryCategory).where(
            InventoryCategory.company_id == company_id,
            InventoryCategory.name == name,
            InventoryCategory.parent_id == current_parent_id
        )
        result = await db.execute(stmt)
        category = result.scalar_one_or_none()

        if not category:
            # Pokud neexistuje, vytvoříme ji pod aktuálním rodičem
            category = InventoryCategory(
                name=name, 
                company_id=company_id, 
                parent_id=current_parent_id
            )
            db.add(category)
            await db.flush() # Musíme flushnout, abychom dostali ID pro další cyklus
        
        # Nastavíme aktuální kategorii jako rodiče pro příští iteraci
        current_parent_id = category.id
        last_category = category

    return last_category

async def get_location_by_name(db: AsyncSession, company_id: int, location_name: str):
    """Najde lokaci podle názvu (case-insensitive)."""
    if not location_name:
        return None
    stmt = select(Location).where(
        Location.company_id == company_id,
        Location.name == location_name.strip()
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


async def get_or_create_supplier(db: AsyncSession, company_id: int, name: str) -> Optional[Supplier]:
    """Najde nebo vytvoří dodavatele podle jména."""
    name = name.strip()
    if not name:
        return None
    stmt = select(Supplier).where(Supplier.company_id == company_id, Supplier.name == name)
    supplier = (await db.execute(stmt)).scalar_one_or_none()
    if not supplier:
        supplier = Supplier(company_id=company_id, name=name)
        db.add(supplier)
        await db.flush()
    return supplier


async def get_or_create_manufacturer(db: AsyncSession, company_id: int, name: str) -> Optional[Manufacturer]:
    """Najde nebo vytvoří výrobce podle jména."""
    name = name.strip()
    if not name:
        return None
    stmt = select(Manufacturer).where(Manufacturer.company_id == company_id, Manufacturer.name == name)
    manufacturer = (await db.execute(stmt)).scalar_one_or_none()
    if not manufacturer:
        manufacturer = Manufacturer(company_id=company_id, name=name)
        db.add(manufacturer)
        await db.flush()
    return manufacturer

# --- ENDPOINTY ---

@router.post("/preview")
async def preview_inventory_excel(
    company_id: int,
    file: UploadFile = File(...),
    _: dict = Depends(require_admin_access)
):
    # ... (tento endpoint zůstává stejný) ...
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Soubor musí být formátu .xlsx")

    contents = await file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True, read_only=True)
    ws = wb.active

    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(cell) if cell is not None else f"Sloupec {i+1}" for i, cell in enumerate(row)]
        break
    
    return {"headers": headers}


@router.post("/upload")
async def import_inventory_from_excel(
    company_id: int,
    file: UploadFile = File(...),
    mapping: str = Form(...),
    db: AsyncSession = Depends(get_db),
    token: dict = Depends(require_admin_access)
):
    """Importuje položky a vytváří stromovou strukturu kategorií."""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Soubor musí být formátu .xlsx")

    try:
        col_map = json.loads(mapping)
    except:
        raise HTTPException(status_code=400, detail="Neplatný formát mapování.")

    contents = await file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    user_id = int(token.get("sub"))

    def get_val(row_data, field_key):
        idx = col_map.get(field_key)
        if idx is not None and isinstance(idx, int) and 0 <= idx < len(row_data):
            return row_data[idx]
        return None

    # Indexy sloupců pro hierarchii kategorií (seznam čísel)
    category_level_indices: List[int] = col_map.get('category_levels', [])

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = get_val(row, 'name')
            sku = str(get_val(row, 'sku')).strip() if get_val(row, 'sku') else None

            if not name or not sku:
                stats["skipped"] += 1
                continue

            # Načtení hodnot
            alt_sku_val = get_val(row, 'alt_sku')
            alt_sku = str(alt_sku_val).strip() if alt_sku_val else None
            ean_val = get_val(row, 'ean')
            ean = str(ean_val).strip() if ean_val else None
            price_val = get_val(row, 'price')
            price = float(price_val) if price_val else 0.0
            desc_val = get_val(row, 'description')
            description = str(desc_val) if desc_val else None
            qty_val = get_val(row, 'quantity')
            quantity = int(qty_val) if qty_val else 0
            loc_val = get_val(row, 'location')
            location_name = str(loc_val).strip() if loc_val else None

            # Dodavatel a výrobce
            supplier_val = get_val(row, 'supplier')
            manufacturer_val = get_val(row, 'manufacturer')
            supplier = await get_or_create_supplier(db, company_id, str(supplier_val)) if supplier_val else None
            manufacturer = await get_or_create_manufacturer(db, company_id, str(manufacturer_val)) if manufacturer_val else None

            # 1. Zpracování hierarchie kategorií ze sloupců
            categories_to_assign = []
            if category_level_indices:
                # Sestavíme cestu z hodnot buněk na vybraných sloupcích (přeskočíme prázdné)
                parts = []
                for col_idx in category_level_indices:
                    if 0 <= col_idx < len(row):
                        cell_val = row[col_idx]
                        if cell_val is not None and str(cell_val).strip():
                            parts.append(str(cell_val).strip())
                if parts:
                    path_str = " > ".join(parts)
                    cat = await resolve_category_path(db, company_id, path_str)
                    if cat:
                        categories_to_assign.append(cat)

            # 2. Položka
            stmt_item = select(InventoryItem).where(
                InventoryItem.company_id == company_id,
                InventoryItem.sku == sku
            ).options(selectinload(InventoryItem.categories))
            item = (await db.execute(stmt_item)).scalar_one_or_none()

            if item:
                # Update
                item.name = str(name)
                if alt_sku:
                    item.alternative_sku = alt_sku
                if ean:
                    item.ean = ean
                if price > 0:
                    item.price = price
                if description:
                    item.description = description
                if supplier:
                    item.supplier_id = supplier.id
                if manufacturer:
                    item.manufacturer_id = manufacturer.id

                # Přidání nových kategorií
                for cat in categories_to_assign:
                    if cat not in item.categories:
                        item.categories.append(cat)

                stats["updated"] += 1
            else:
                # Create
                item = InventoryItem(
                    company_id=company_id,
                    name=str(name),
                    sku=sku,
                    alternative_sku=alt_sku,
                    ean=ean,
                    price=price,
                    description=description,
                    supplier_id=supplier.id if supplier else None,
                    manufacturer_id=manufacturer.id if manufacturer else None,
                    is_monitored_for_stock=True,
                    low_stock_threshold=5
                )
                item.categories = categories_to_assign

                db.add(item)
                await db.flush()
                stats["created"] += 1

            # 3. Naskladnění
            if quantity > 0 and location_name:
                location = await get_location_by_name(db, company_id, location_name)
                if location:
                    stock = await db.get(ItemLocationStock, (item.id, location.id))
                    old_qty = 0
                    if stock:
                        old_qty = stock.quantity
                        stock.quantity += quantity
                    else:
                        stock = ItemLocationStock(inventory_item_id=item.id, location_id=location.id, quantity=quantity)
                        db.add(stock)
                    
                    log = InventoryAuditLog(
                        item_id=item.id, user_id=user_id, company_id=company_id,
                        action=AuditLogAction.quantity_adjusted,
                        details=f"Import XLS: +{quantity} ks na '{location.name}'. (Původně: {old_qty})"
                    )
                    db.add(log)
                else:
                    stats["errors"].append(f"Řádek {index}: Lokace '{location_name}' nenalezena.")

        except Exception as e:
            stats["errors"].append(f"Řádek {index}: Chyba - {str(e)}")
            continue

    await db.commit()
    return stats


@router.get("/export")
async def export_inventory_excel(
    company_id: int,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(require_admin_access)
):
    """Vyexportuje kompletní stav skladu včetně stromové struktury kategorií."""
    
    # 1. Načtení všech kategorií do paměti pro rychlé sestavení stromu
    cat_stmt = select(InventoryCategory).where(InventoryCategory.company_id == company_id)
    all_cats = (await db.execute(cat_stmt)).scalars().all()
    cat_map = {c.id: c for c in all_cats}

    def get_full_path(category_id):
        """Rekurzivně sestaví cestu: Hlavní > Podkategorie > Podkategorie"""
        path = []
        curr = cat_map.get(category_id)
        while curr:
            path.append(curr.name)
            curr = cat_map.get(curr.parent_id) if curr.parent_id else None
        return " > ".join(reversed(path))

    # 2. Načtení položek
    stmt = (
        select(InventoryItem)
        .where(InventoryItem.company_id == company_id)
        .options(
            selectinload(InventoryItem.categories),
            selectinload(InventoryItem.locations).selectinload(ItemLocationStock.location),
            selectinload(InventoryItem.manufacturer),
            selectinload(InventoryItem.supplier)
        )
        .order_by(InventoryItem.name)
    )
    result = await db.execute(stmt)
    items = result.scalars().all()

    # 3. Vytvoření Excelu
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sklad"

    headers = [
        "Název položky", "SKU", "Alt. SKU", "EAN", 
        "Kategorie (Strom)", 
        "Výrobce", "Dodavatel", "Nákupní cena", "Prodejní cena (MOC)", 
        "DPH %", "Celkem ks", "Hlídat stav", "Min. limit", 
        "Popis", "Lokace (Detail)"
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 4. Naplnění daty
    for item in items:
        # Sestavení stromové cesty pro všechny kategorie (oddělené novým řádkem)
        cat_paths = [get_full_path(c.id) for c in item.categories]
        cat_str = "\n".join(cat_paths)
        
        loc_details = ", ".join([f"{l.location.name}: {l.quantity}" for l in item.locations if l.quantity > 0])
        total_qty = sum(loc.quantity for loc in item.locations)

        row = [
            item.name,
            item.sku,
            item.alternative_sku or "",
            item.ean or "",
            cat_str,
            item.manufacturer.name if item.manufacturer else "",
            item.supplier.name if item.supplier else "",
            item.price,
            item.retail_price,
            item.vat_rate,
            total_qty,
            "ANO" if item.is_monitored_for_stock else "NE",
            item.low_stock_threshold if item.is_monitored_for_stock else "",
            item.description,
            loc_details
        ]
        ws.append(row)

    # 5. Formátování
    col_widths = {
        'A': 30, 'B': 15, 'C': 15, 'D': 15,
        'E': 40, # Kategorie
        'F': 20, 'G': 20,
        'N': 30, 'O': 50
    }
    
    for col_char, width in col_widths.items():
        ws.column_dimensions[col_char].width = width

    # Zalamování textu pro kategorie a lokace
    for row in ws.iter_rows(min_row=2, max_col=15):
        # Sloupec E (index 4 v poli, ale v Excelu 5) a O (15)
        row[4].alignment = Alignment(wrap_text=True, vertical='top') 
        row[14].alignment = Alignment(wrap_text=True, vertical='top')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"sklad_komplet_{company_id}.xlsx"
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )