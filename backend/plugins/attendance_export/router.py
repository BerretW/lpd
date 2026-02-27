from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from datetime import date, datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import locale

# Nastavení české lokalizace pro názvy měsíců
MONTH_NAMES = ["", "Leden", "Únor", "Březen", "Duben", "Květen", "Červen", 
               "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"]

from app.db.database import get_db
# --- ZMĚNA ZDE: Přidány importy User, Task, WorkOrder ---
from app.db.models import TimeLog, TimeLogEntryType, User, Task, WorkOrder
from app.core.dependencies import require_company_access

router = APIRouter(prefix="/plugins/attendance-export", tags=["plugin-attendance-export"])

# --- POMOCNÉ FUNKCE PRO STYLOVÁNÍ ---

def apply_border(cell, style='thin'):
    side = Side(border_style=style, color="000000")
    cell.border = Border(left=side, right=side, top=side, bottom=side)

def set_column_widths(ws):
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 8
    ws.column_dimensions['L'].width = 8
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 8
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 5
    ws.column_dimensions['Q'].width = 8
    ws.column_dimensions['R'].width = 15
    ws.column_dimensions['S'].width = 15

# --- GENEROVÁNÍ MĚSÍČNÍHO LISTU ---

def create_monthly_sheet(wb, month_index, year, logs, user_name):
    month_name = MONTH_NAMES[month_index]
    ws = wb.create_sheet(title=f"Měsíční plán {month_name}")
    set_column_widths(ws)

    # Hlavička
    ws.merge_cells('A1:S1')
    ws['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    ws.merge_cells('A2:C2')
    ws['A2'] = month_name
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('F2:H2')
    ws['F2'] = year
    ws['F2'].font = Font(bold=True, size=14)
    ws['F2'].alignment = Alignment(horizontal='center')

    ws['K2'] = "Jméno:"
    ws['K2'].font = Font(bold=True)
    ws.merge_cells('L2:N2')
    ws['L2'] = user_name
    ws['L2'].font = Font(bold=True, size=12)

    ws.merge_cells('R2:S2')
    ws['R2'] = "Překážka v práci"
    ws['R2'].font = Font(bold=True)
    ws['R2'].alignment = Alignment(horizontal='center')

    headers = [
        ("Datum", "A3"), ("Příjezd", "B3"), ("Odjezd", "C3"), ("Město", "D3"), 
        ("Instituce", "E3"), ("Km", "F3"), ("Čas", "G3"), ("Akce", "H3"), 
        ("Auto", "I3"), ("Litry", "J3"), ("Kč", "K3"), ("Služba", "L3"),
        ("V práci od-do", "M3"), ("Oběd", "N3"), ("Dovolená", "O3"), ("NV", "P3"),
        ("Výjezd", "Q3"), ("Zaměstnavat", "R3"), ("Zaměstnane", "S3")
    ]

    for title, cell_ref in headers:
        cell = ws[cell_ref]
        cell.value = title
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell, 'medium')

    # Data
    from calendar import monthrange
    days_in_month = monthrange(year, month_index)[1]
    
    current_row = 4
    
    for day in range(1, days_in_month + 1):
        current_date = date(year, month_index, day)
        date_str = current_date.strftime("%d.%m.%Y")
        
        daily_logs = [l for l in logs if l.start_time.date() == current_date]
        daily_logs.sort(key=lambda x: x.start_time)

        loops = max(len(daily_logs), 1)
        
        if loops > 1:
            ws.merge_cells(f'A{current_row}:A{current_row + loops - 1}')
        
        ws[f'A{current_row}'].value = date_str
        ws[f'A{current_row}'].alignment = Alignment(vertical='top')
        
        is_weekend = current_date.weekday() >= 5
        fill_color = "DDDDDD" if is_weekend else None

        for i in range(loops):
            row_idx = current_row + i
            log = daily_logs[i] if i < len(daily_logs) else None
            
            if fill_color:
                for col in range(1, 20):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            for col in range(1, 20):
                apply_border(ws.cell(row=row_idx, column=col))

            if log:
                start_time_str = log.start_time.strftime("%H:%M")
                end_time_str = log.end_time.strftime("%H:%M")
                
                duration = (log.end_time - log.start_time).total_seconds() / 3600
                
                ws[f'B{row_idx}'] = start_time_str
                ws[f'C{row_idx}'] = end_time_str
                
                client_name = ""
                city = ""
                if log.task and log.task.work_order and log.task.work_order.client:
                    client_name = log.task.work_order.client.name
                    if log.task.work_order.client.address:
                        parts = log.task.work_order.client.address.split(',')
                        city = parts[-1].strip() if len(parts) > 0 else log.task.work_order.client.address

                ws[f'D{row_idx}'] = city
                ws[f'E{row_idx}'] = client_name
                ws[f'H{row_idx}'] = log.task.name if log.task else (log.notes or "")

                if log.entry_type == TimeLogEntryType.WORK:
                    ws[f'M{row_idx}'] = f"{start_time_str}\n{end_time_str}"
                    ws[f'M{row_idx}'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    if log.break_duration_minutes > 0:
                        ws[f'N{row_idx}'] = f"{log.break_duration_minutes} min"

                elif log.entry_type == TimeLogEntryType.VACATION:
                    ws[f'O{row_idx}'] = 8
                
                elif log.entry_type == TimeLogEntryType.SICK_DAY:
                    ws[f'S{row_idx}'] = 8

                elif log.entry_type == TimeLogEntryType.DOCTOR:
                    ws[f'S{row_idx}'] = duration

        current_row += loops

# --- GENEROVÁNÍ ROČNÍHO SOUHRNU ---

def create_yearly_summary(wb, year, all_logs, user_name):
    ws = wb.create_sheet(title="Výkaz práce Roční")
    wb.active = ws
    
    widths = [20] + [12] * 12
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    ws['A1'] = f"Pracovní výkaz za rok: {year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['E1'] = f"Jméno pracovníka: {user_name}"
    ws['E1'].font = Font(bold=True, size=12)

    headers = [
        "Měsíc", "Odprac. celkem", "Z toho přesčas", "Překážka Zaměstnavatel", 
        "Překážka Zaměstnanec", "Dovolená", "Hodin v noci", "Hodin celkem",
        "Náhradní volno", "Pohotovost", "Zbývá dovolené", "K proplacení", "Fond hodin"
    ]
    
    ws.append([])
    ws.append([]) 
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        apply_border(cell, 'medium')
        if "Fond" in header:
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    row_idx = 5
    total_year_hours = 0
    
    for month in range(1, 13):
        month_logs = [l for l in all_logs if l.start_time.month == month]
        
        work_hours = sum((l.end_time - l.start_time).total_seconds() / 3600 for l in month_logs if l.entry_type == TimeLogEntryType.WORK)
        vacation_hours = sum((l.end_time - l.start_time).total_seconds() / 3600 for l in month_logs if l.entry_type == TimeLogEntryType.VACATION)
        sick_hours = sum((l.end_time - l.start_time).total_seconds() / 3600 for l in month_logs if l.entry_type in [TimeLogEntryType.SICK_DAY, TimeLogEntryType.DOCTOR])
        overtime_hours = sum((l.end_time - l.start_time).total_seconds() / 3600 for l in month_logs if getattr(l, 'is_overtime', False))
        
        total_hours = work_hours + vacation_hours + sick_hours
        total_year_hours += total_hours

        row_data = [
            MONTH_NAMES[month],
            round(work_hours, 2),
            round(overtime_hours, 2),
            0,
            round(sick_hours, 2),
            round(vacation_hours, 2),
            0,
            round(total_hours, 2),
            0,
            0,
            0,
            round(total_hours, 2),
            168
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_border(cell)
            if col_idx > 1:
                cell.number_format = '0.00'
        
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="CELKEM").font = Font(bold=True)
    ws.cell(row=row_idx, column=8, value=round(total_year_hours, 2)).font = Font(bold=True)


# --- HLAVNÍ ENDPOINT ---

@router.get("/download")
async def download_attendance_excel(
    company_id: int,
    year: int,
    user_id: int = None,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_company_access)
):
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)

    # --- ZMĚNA ZDE: POUŽITÍ TŘÍD MÍSTO STRINGŮ V selectinload ---
    stmt = (
        select(TimeLog)
        .where(
            TimeLog.company_id == company_id,
            TimeLog.start_time >= start_date,
            TimeLog.start_time <= f"{end_date} 23:59:59"
        )
        .options(
            selectinload(TimeLog.user),
            # Řetězení: TimeLog.task -> Task.work_order -> WorkOrder.client
            selectinload(TimeLog.task).selectinload(Task.work_order).selectinload(WorkOrder.client),
            selectinload(TimeLog.work_type)
        )
        .order_by(TimeLog.start_time)
    )

    if user_id:
        stmt = stmt.where(TimeLog.user_id == user_id)

    result = await db.execute(stmt)
    all_logs = result.scalars().all()

    if not all_logs and user_id:
        user = await db.get(User, user_id)
        user_name = user.email if user else "Neznámý"
    elif all_logs:
        user_name = all_logs[0].user.email
    else:
        user_name = "Neznámý"

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    create_yearly_summary(wb, year, all_logs, user_name)

    for month in range(1, 13):
        month_logs = [l for l in all_logs if l.start_time.month == month]
        create_monthly_sheet(wb, month, year, month_logs, user_name)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"vykaz_{year}_{user_name}.xlsx"
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )